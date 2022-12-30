
import datetime
from datetime import datetime, timedelta
from time import sleep

import instaloader
import yaml
from os import path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from itertools import takewhile

#reads the config.yaml file
def read(config_file):
    """
    Read yaml config file from current directory
    """
    print('Reading', config_file)
    with open(config_file, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    return config


def save(config_file,data):
    """
    Read yaml config file from current directory
    """
    print('Writing', config_file)
    with open(config_file, 'w', encoding='utf-8') as f:
        yaml.dump(data,f,default_flow_style=False)

def setupconfig():
    config = read("config.yaml")
    TrackedNames = config.get('TrackedNames',[])
    TrackLifetime = config.get('TrackLifetime')
    currenttracked = read("currenttracked.yaml")
    username = config.get('LoginUsername')
    password = config.get('LoginPassword')
    return TrackedNames, TrackLifetime, currenttracked, username, password


def setupworkbook(workbook_name):
    workbook_name = f'Results/{workbook_name}'
    if path.exists(workbook_name):
        wb = load_workbook(workbook_name)
        page = wb.active
    else:
        headers = ['Profile Name', 'Current date', 'URL', 'Post Type', 'Date Posted', 'Current Likes',
                   'Current Comments',
                   'Current View']  # ,'Video duration','Is Sponsored?'] ##'Current View','Video duration'
        wb = Workbook()
        page = wb.active
        page.title = 'Data'
        page.append(headers)
        wb.save(filename=workbook_name)
    return wb, page

#def GetData(bot):



def main ():

    TrackedNames, TrackLifetime, currenttracked, username, password = setupconfig()
    date = datetime.today() - timedelta(hours=TrackLifetime+8)

    print (f"tracking {TrackedNames} for {TrackLifetime} hours")

    print("initialize instaloader bot")
    bot = instaloader.Instaloader()

    if path.exists('session'):
        bot.load_session_from_file(username,'session')
    if bot.test_login() == None:
        bot.login(user=username, passwd=password)
        print (bot.test_login())
    #get the data
    for name in TrackedNames:

        wb, page = setupworkbook(f'{name}.xlsx')
        print(f"Results will be stored in {name}.xlsx")


        profile = instaloader.Profile.from_username(bot.context,name)
        posts = profile.get_posts()
        for post in takewhile(lambda p: p.date > date, posts):
            postsh = str(post.shortcode)
            #checks if we should get the data from the post or not
            if postsh not in currenttracked:
                currenttracked[postsh] = TrackLifetime
            elif postsh in currenttracked:
                if currenttracked[postsh] == 0:
                    continue

            try:
                data = [post.owner_username, datetime.now(), "https://www.instagram.com/p/" + postsh + "/",
                        post.typename, (post.date - timedelta(hours=-8)), post.likes, post.comments,
                        post.video_view_count]  # ,Post.video_duration]#,Post.is_sponsored]#Post.video_view_count,Post.video_duration,
                page.append(data)

            except:
                print("error. Instagram blocked the connection!")
                exit()

            if currenttracked[postsh] >= 1:
                currenttracked[postsh] -= 1
                save('currenttracked.yaml', currenttracked)
        wb.save(filename=f'Results/{name}.xlsx')
        wb.close()
    bot.save_session_to_file('session')






















# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
